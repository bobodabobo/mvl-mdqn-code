import numpy as np
import pickle
from openpyxl import load_workbook

from .keys import systems, tasks, DQN_methods, heuristic_methods


def createTable():
    datas = _read_data()
    _write(datas)


def _read_data():
    # read DQN results
    DQN_file_name = "results/DQN_results.pkl"
    with open(DQN_file_name, "rb") as f:
        DQN_results = pickle.load(f)
    n_seeds = len(DQN_results['LS']['1']['DQN'])
    # read heuristic results
    heuristic_file_name = "results/heuristic_results.pkl"
    with open(heuristic_file_name, "rb") as f:
        heuristic_results = pickle.load(f)
    # init data matrix
    datas = np.zeros((len(systems) * (len(DQN_methods) + 2), len(tasks)), dtype=np.float64)
    # read data
    for i, inv_sys in enumerate(systems):
        for j, task in enumerate(tasks):
            # DQN methods
            for k, DQN_method in enumerate(DQN_methods):
                DQN_results_all_seeds = sorted([DQN_results[inv_sys][task][DQN_method][s]['performance'] for s in range(n_seeds)])
                datas[i * (len(DQN_methods) + 2) + k, j] = DQN_results_all_seeds[0] # best performanc
                # datas[i * (len(DQN_methods) + 2) + k, j] = np.mean(DQN_results_all_seeds[:4]) # mean of top 1/4 performances
            # heuristic methods
            for k, heuristic_method in enumerate(heuristic_methods[i]):
                datas[i * (len(DQN_methods) + 2) + len(DQN_methods) + k, j] = heuristic_results[inv_sys][task][heuristic_method]['performance']
    return datas


def _write(data_mat:np.ndarray):
    templete = load_workbook('utilize/table_templete.xlsx')
    sheet = templete['Sheet1']
    start_row = 2
    start_col = 3
    for i in range(data_mat.shape[0]):
        for j in range(data_mat.shape[1]):
            sheet.cell(row=start_row + i, column=start_col + j, value=data_mat[i, j])
    templete.save('results/table.xlsx')


if __name__ == "__main__":
    createTable()